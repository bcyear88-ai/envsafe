import io
from flask import Blueprint, request, jsonify, send_file
from models import db, WorkPlan, Hazard, Inspection
from routes.utils import login_required
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

report_bp = Blueprint('report', __name__)


@report_bp.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    today = date.today()
    year = request.args.get('year', today.year, type=int)
    month = request.args.get('month', today.month, type=int)

    month_plans = WorkPlan.query.filter(WorkPlan.plan_year == year, WorkPlan.plan_month == month).all()
    year_plans = WorkPlan.query.filter(WorkPlan.plan_year == year).all()
    all_hazards = Hazard.query.all()

    month_total = len(month_plans)
    month_completed = sum(1 for p in month_plans if p.status == 'completed')
    month_in_progress = sum(1 for p in month_plans if p.status == 'in_progress')
    month_pending = sum(1 for p in month_plans if p.status == 'pending')
    month_overdue = 0
    for p in month_plans:
        if p.status != 'completed' and p.deadline and p.deadline < today:
            month_overdue += 1

    year_total = len(year_plans)
    year_completed = sum(1 for p in year_plans if p.status == 'completed')

    hazard_total = len(all_hazards)
    hazard_done = sum(1 for h in all_hazards if h.status in ['rectified', 'closed'])
    hazard_overdue = 0
    for h in all_hazards:
        if h.status not in ['rectified', 'closed'] and h.deadline and h.deadline < today:
            hazard_overdue += 1
    hazard_pending = sum(1 for h in all_hazards if h.status == 'pending')

    soon_expire = []
    for p in month_plans:
        if p.status != 'completed' and p.deadline:
            days_left = (p.deadline - today).days
            if 0 <= days_left <= 7:
                soon_expire.append({'type': '计划', 'title': p.title, 'deadline': str(p.deadline), 'days_left': days_left})
    for h in all_hazards:
        if h.status not in ['rectified', 'closed'] and h.deadline:
            days_left = (h.deadline - today).days
            if 0 <= days_left <= 7:
                soon_expire.append({'type': '隐患', 'title': h.description or h.title or '未命名',
                                    'deadline': str(h.deadline), 'days_left': days_left})

    month_complete_rate = round(month_completed / month_total * 100, 1) if month_total else 0
    hazard_complete_rate = round(hazard_done / hazard_total * 100, 1) if hazard_total else 0

    trend = []
    for m in range(1, 13):
        mp = WorkPlan.query.filter(WorkPlan.plan_year == year, WorkPlan.plan_month == m).all()
        t = len(mp)
        c = sum(1 for p in mp if p.status == 'completed')
        hr = 0
        month_hazards = [h for h in all_hazards if h.check_date and h.check_date.year == year and h.check_date.month == m]
        ht = len(month_hazards)
        hd = sum(1 for h in month_hazards if h.status in ['rectified', 'closed'])
        hr = round(hd / ht * 100, 1) if ht else 0
        trend.append({'month': m, 'plan_total': t, 'plan_completed': c,
                      'plan_rate': round(c / t * 100, 1) if t else 0,
                      'hazard_total': ht, 'hazard_done': hd, 'hazard_rate': hr})

    category_stats = []
    cat_map = {}
    for p in year_plans:
        cat = p.category or '未分类'
        if cat not in cat_map:
            cat_map[cat] = {'name': cat, 'total': 0, 'done': 0}
        cat_map[cat]['total'] += 1
        if p.status == 'completed':
            cat_map[cat]['done'] += 1
    for v in cat_map.values():
        v['rate'] = round(v['done'] / v['total'] * 100, 1) if v['total'] else 0
    category_stats = sorted(cat_map.values(), key=lambda x: -x['rate'])

    hazard_status = []
    hs_map = {'closed': 0, 'rectified': 0, 'rectifying': 0, 'pending': 0}
    for h in all_hazards:
        hs_map[h.status] = hs_map.get(h.status, 0) + 1
    hazard_status = [
        {'name': '已闭环', 'value': hs_map.get('closed', 0)},
        {'name': '待验收', 'value': hs_map.get('rectified', 0)},
        {'name': '整改中', 'value': hs_map.get('rectifying', 0)},
        {'name': '待整改', 'value': hs_map.get('pending', 0)},
    ]

    major_count = sum(1 for h in all_hazards if h.level == 'major')
    general_count = sum(1 for h in all_hazards if h.level == 'general')

    return jsonify({'code': 0, 'data': {
        'month': {'total': month_total, 'completed': month_completed, 'in_progress': month_in_progress,
                  'pending': month_pending, 'overdue': month_overdue, 'complete_rate': month_complete_rate},
        'year': {'total': year_total, 'completed': year_completed,
                 'complete_rate': round(year_completed / year_total * 100, 1) if year_total else 0},
        'hazard': {'total': hazard_total, 'done': hazard_done, 'pending': hazard_pending,
                   'overdue': hazard_overdue, 'complete_rate': hazard_complete_rate,
                   'major_count': major_count, 'general_count': general_count},
        'soon_expire': soon_expire[:10],
        'trend': trend,
        'category_stats': category_stats,
        'hazard_status': hazard_status,
        'inspection_count': Inspection.query.count(),
    }})


@report_bp.route('/export', methods=['GET'])
@login_required
def export_report():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center')

    ws_plan = wb.active
    ws_plan.title = '工作计划'
    ws_plan.append(['序号', '标题', '类别', '类型', '年份', '月份', '负责人', '截止日期', '进度%', '状态', '完成日期', '法规依据'])
    for cell in ws_plan[1]:
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.alignment = header_align

    plans = WorkPlan.query.filter(WorkPlan.plan_year == year).order_by(WorkPlan.plan_month).all()
    if month:
        plans = [p for p in plans if p.plan_month == month]

    status_map = {'pending': '待开始', 'in_progress': '进行中', 'completed': '已完成', 'overdue': '已超期'}
    for i, p in enumerate(plans, 1):
        ws_plan.append([i, p.title, p.category, p.type or '', p.plan_year, p.plan_month or '',
                        p.responsible or '', p.deadline or '', p.progress, status_map.get(p.status, p.status),
                        p.completion_date or '', p.law_basis or ''])
    for col in ws_plan.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws_plan.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws_hazard = wb.create_sheet('隐患整改')
    ws_hazard.append(['序号', '隐患描述', '级别', '位置', '来源', '负责人', '期限', '整改进度%', '状态', '整改措施', '验收人', '验收结果', '法规依据'])
    for cell in ws_hazard[1]:
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='ED7D31')
        cell.alignment = header_align

    hazards = Hazard.query.order_by(Hazard.created_at.desc()).all()
    level_map = {'major': '重大', 'general': '一般'}
    hstatus_map = {'pending': '待整改', 'rectifying': '整改中', 'rectified': '待验收', 'closed': '已闭环'}
    for i, h in enumerate(hazards, 1):
        ws_hazard.append([i, (h.description or h.title or '')[:80],
                          level_map.get(h.level, h.level), h.location or '', h.source or '',
                          h.responsible or '', h.deadline or '', h.rectify_progress or 0,
                          hstatus_map.get(h.status, h.status),
                          (h.rectify_measure or '')[:80], h.acceptor or '', h.accept_result or '', h.law_basis or ''])
    for col in ws_hazard.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws_hazard.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws_summary = wb.create_sheet('统计汇总')
    ws_summary.append(['{}年{}工作完成情况汇总'.format(year, '{}月'.format(month) if month else '全年')])
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(['指标', '数值'])
    ws_summary[3][0].font = Font(bold=True)

    total_plan = len(plans)
    completed_plan = sum(1 for p in plans if p.status == 'completed')
    ws_summary.append(['工作计划总数', total_plan])
    ws_summary.append(['已完成计划数', completed_plan])
    ws_summary.append(['计划完成率', '{}%'.format(round(completed_plan / total_plan * 100, 1) if total_plan else 0)])

    total_hazard = len(hazards)
    done_hazard = sum(1 for h in hazards if h.status in ['rectified', 'closed'])
    ws_summary.append([])
    ws_summary.append(['隐患总数', total_hazard])
    ws_summary.append(['已整改/闭环', done_hazard])
    ws_summary.append(['隐患整改率', '{}%'.format(round(done_hazard / total_hazard * 100, 1) if total_hazard else 0)])
    ws_summary.append(['重大隐患', sum(1 for h in hazards if h.level == 'major')])
    ws_summary.append(['一般隐患', sum(1 for h in hazards if h.level == 'general')])

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = '安环工作报表_{}年{}.xlsx'.format(year, '{}月'.format(month) if month else '全年')
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
